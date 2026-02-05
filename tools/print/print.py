"""Tisk nových kartiček: Excel → HTML + zápis printed_at.

Spuštění z kořene projektu:
    python tools/print/print.py
"""

import json
import os
import sys
from datetime import datetime

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, "..", ".."))

ALL_COLUMNS = ["en", "pron", "meaning_en", "cz", "rating", "printed_at", "note"]


def load_configs():
    with open(os.path.join(ROOT_DIR, "tools", "config.json"), encoding="utf-8") as f:
        main_cfg = json.load(f)
    with open(os.path.join(SCRIPT_DIR, "config.json"), encoding="utf-8") as f:
        print_cfg = json.load(f)
    return main_cfg, print_cfg


def read_unprinted(excel_path):
    """Read rows where printed_at is empty. Returns (wb, ws, col_map, rows_to_print).
    rows_to_print = list of (row_index_1based, {col: value})
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    header = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    col_map = {name: idx for idx, name in enumerate(header) if name in ALL_COLUMNS}

    if "en" not in col_map:
        print("ERROR: Missing 'en' column.")
        sys.exit(1)

    printed_at_col = col_map.get("printed_at")
    rows = []

    for row_idx in range(2, ws.max_row + 1):
        en_val = ws.cell(row=row_idx, column=col_map["en"] + 1).value
        if not en_val or not str(en_val).strip():
            continue

        if printed_at_col is not None:
            pa_val = ws.cell(row=row_idx, column=printed_at_col + 1).value
            if pa_val is not None and str(pa_val).strip():
                continue

        word = {}
        for col_name, col_idx in col_map.items():
            val = ws.cell(row=row_idx, column=col_idx + 1).value
            word[col_name] = str(val).strip() if val is not None else ""
        rows.append((row_idx, word))

    return wb, ws, col_map, rows


def generate_html(rows, print_cfg):
    front_fields = print_cfg.get("front", ["en", "pron"])
    back_fields = print_cfg.get("back", ["meaning_en", "cz"])

    cards_html = []
    for _, word in rows:
        front_lines = [word.get(f, "") for f in front_fields]
        back_lines = [word.get(f, "") for f in back_fields]

        front_main = front_lines[0] if front_lines else ""
        front_sub = "<br>".join(front_lines[1:]) if len(front_lines) > 1 else ""
        back_main = back_lines[0] if back_lines else ""
        back_sub = "<br>".join(back_lines[1:]) if len(back_lines) > 1 else ""

        cards_html.append(f"""      <div class="card">
        <div class="front">
          <div class="main">{front_main}</div>
          <div class="sub">{front_sub}</div>
        </div>
        <div class="divider"></div>
        <div class="back">
          <div class="main">{back_main}</div>
          <div class="sub">{back_sub}</div>
        </div>
      </div>""")

    grid_blocks = []
    for i in range(0, len(cards_html), 8):
        page_cards = "\n".join(cards_html[i : i + 8])
        grid_blocks.append(f'    <div class="page">\n{page_cards}\n    </div>')

    pages = "\n".join(grid_blocks)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <title>Flashcards Print</title>
  <style>
    * {{ margin: 0; padding: 0; box-sizing: border-box; }}
    @page {{ size: A4; margin: 10mm; }}

    .page {{
      width: 190mm;
      height: 277mm;
      display: grid;
      grid-template-columns: 1fr 1fr;
      grid-template-rows: repeat(4, 1fr);
      gap: 2mm;
      page-break-after: always;
    }}

    .card {{
      border: 1px solid #ccc;
      display: flex;
      flex-direction: column;
      font-family: Arial, sans-serif;
    }}

    .front, .back {{
      flex: 1;
      display: flex;
      flex-direction: column;
      justify-content: center;
      align-items: center;
      padding: 4mm;
      text-align: center;
    }}

    .front .main {{
      font-size: 16pt;
      font-weight: bold;
    }}

    .front .sub {{
      font-size: 11pt;
      color: #666;
      margin-top: 2mm;
    }}

    .back .main {{
      font-size: 13pt;
    }}

    .back .sub {{
      font-size: 11pt;
      color: #555;
      margin-top: 2mm;
    }}

    .divider {{
      border-top: 1px dashed #999;
      height: 0;
    }}

    @media print {{
      body {{ margin: 0; }}
      .page {{ page-break-after: always; }}
      .page:last-child {{ page-break-after: auto; }}
    }}
  </style>
</head>
<body>
{pages}
</body>
</html>
"""


def stamp_printed_at(wb, ws, col_map, rows):
    pa_col_idx = col_map.get("printed_at")
    if pa_col_idx is None:
        print("WARNING: No 'printed_at' column found, skipping stamp.")
        return

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for row_idx, _ in rows:
        ws.cell(row=row_idx, column=pa_col_idx + 1, value=now)


def main():
    os.chdir(ROOT_DIR)

    main_cfg, print_cfg = load_configs()
    excel_path = main_cfg["excel_path"]

    if not os.path.exists(excel_path):
        print(f"ERROR: Excel not found: {excel_path}")
        sys.exit(1)

    print(f"Reading: {excel_path}")
    wb, ws, col_map, rows = read_unprinted(excel_path)

    if not rows:
        print("No new (unprinted) cards found.")
        wb.close()
        return

    print(f"Found {len(rows)} unprinted cards.")

    html = generate_html(rows, print_cfg)
    output_path = os.path.join("tools", "print", "output.html")
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Written: {output_path}")

    stamp_printed_at(wb, ws, col_map, rows)
    wb.save(excel_path)
    print(f"Stamped printed_at for {len(rows)} rows in {excel_path}.")
    wb.close()

    print("Print done. Open output.html in browser and press Ctrl+P.")


if __name__ == "__main__":
    main()
